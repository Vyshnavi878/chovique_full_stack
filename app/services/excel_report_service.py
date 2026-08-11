import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportService:

    @staticmethod
    def _apply_styles(ws, max_col):
        # Grid lines
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="1a1512")
        meta_font = Font(name=font_family, size=10, italic=True, color="555555")
        section_font = Font(name=font_family, size=12, bold=True, color="c9a84c")
        header_font = Font(name=font_family, size=11, bold=True, color="ffffff")
        data_font = Font(name=font_family, size=11, color="1a1512")
        total_font = Font(name=font_family, size=11, bold=True, color="1a1512")

        header_fill = PatternFill(start_color="1a1512", end_color="1a1512", fill_type="solid")
        total_fill = PatternFill(start_color="f9f6f0", end_color="f9f6f0", fill_type="solid")
        kpi_fill = PatternFill(start_color="fbf9f6", end_color="fbf9f6", fill_type="solid")

        thin_side = Side(border_style="thin", color="dcdcdc")
        double_side = Side(border_style="double", color="1a1512")
        thick_side = Side(border_style="medium", color="c9a84c")

        border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thin_side, bottom=double_side)
        border_kpi = Border(left=thin_side, right=thin_side, top=thick_side, bottom=thin_side)

        return (
            title_font, meta_font, section_font, header_font, data_font, total_font,
            header_fill, total_fill, kpi_fill, border_data, border_total, border_kpi
        )

    @staticmethod
    def _auto_adjust_columns(ws):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                if cell.number_format and ('₹' in cell.number_format or '%' in cell.number_format):
                    max_len = max(max_len, len(val) + 6)
                else:
                    max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @classmethod
    def generate_customer_report(cls, start_date: str, end_date: str, kpis: list, customers: list) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        (
            title_font, meta_font, section_font, header_font, data_font, total_font,
            header_fill, total_fill, kpi_fill, border_data, border_total, border_kpi
        ) = cls._apply_styles(ws, 6)

        # Title
        ws["A1"] = "CHOVIQUE LUXURY CHOCOLATES — CUSTOMERS REPORT"
        ws["A1"].font = title_font
        ws.merge_cells("A1:F1")

        # Date Range
        ws["A2"] = f"Date Range: {start_date} to {end_date}"
        ws["A2"].font = meta_font
        ws.merge_cells("A2:F2")

        # KPI Header
        ws["A4"] = "KPI SUMMARY"
        ws["A4"].font = section_font

        # KPIs Layout
        kpi_cols = ["A", "B", "C", "D", "E"]
        for idx, kpi in enumerate(kpis):
            col = kpi_cols[idx]
            val_cell = ws[f"{col}5"]
            lbl_cell = ws[f"{col}6"]

            lbl_cell.value = kpi.title
            lbl_cell.font = Font(name="Segoe UI", size=9, bold=True, color="555555")
            lbl_cell.alignment = Alignment(horizontal="center")
            lbl_cell.fill = kpi_fill
            lbl_cell.border = border_data

            val_cell.value = kpi.value
            val_cell.font = Font(name="Segoe UI", size=12, bold=True, color="c9a84c")
            val_cell.alignment = Alignment(horizontal="center")
            val_cell.fill = kpi_fill
            val_cell.border = border_kpi

            # Try formatting
            if "%" in str(kpi.value):
                try:
                    val_cell.value = float(str(kpi.value).replace("%", "")) / 100.0
                    val_cell.number_format = "0.0%"
                except:
                    pass
            elif "₹" in str(kpi.value):
                try:
                    val_cell.value = float(str(kpi.value).replace("₹", "").replace(",", ""))
                    val_cell.number_format = "₹#,##0.00"
                except:
                    pass
            else:
                try:
                    val_cell.value = int(str(kpi.value).replace(",", ""))
                except:
                    pass

        # Detailed Table Header
        ws["A8"] = "DETAILED CUSTOMER DATA"
        ws["A8"].font = section_font

        headers = ["Customer Name", "Email", "Phone", "Orders Placed", "Total Spend", "Joined Date"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left" if col_idx <= 3 else "right")

        # Freeze panes at A10
        ws.freeze_panes = "A10"

        # Data rows
        start_row = 10
        for r_idx, cust in enumerate(customers, start_row):
            # Name, Email, Phone
            ws.cell(row=r_idx, column=1, value=cust[0]).font = data_font
            ws.cell(row=r_idx, column=2, value=cust[1]).font = data_font
            ws.cell(row=r_idx, column=3, value=cust[2]).font = data_font

            # Orders Placed (int)
            orders_cell = ws.cell(row=r_idx, column=4, value=int(cust[3] or 0))
            orders_cell.font = data_font
            orders_cell.alignment = Alignment(horizontal="right")
            orders_cell.number_format = "#,##0"

            # Total Spend (float)
            spend_val = float(str(cust[4]).replace("₹", "").replace(",", "")) if cust[4] else 0.0
            spend_cell = ws.cell(row=r_idx, column=5, value=spend_val)
            spend_cell.font = data_font
            spend_cell.alignment = Alignment(horizontal="right")
            spend_cell.number_format = "₹#,##0.00"

            # Joined Date
            date_cell = ws.cell(row=r_idx, column=6)
            try:
                date_cell.value = datetime.strptime(cust[5], "%Y-%m-%d")
                date_cell.number_format = "YYYY-MM-DD"
            except:
                date_cell.value = cust[5]
            date_cell.font = data_font
            date_cell.alignment = Alignment(horizontal="right")

            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).border = border_data

        end_row = start_row + len(customers) - 1
        total_row = end_row + 1

        # Totals Row
        ws.cell(row=total_row, column=1, value="Total Customers").font = total_font
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=total_row, column=2, value=f"=COUNTA(A10:A{end_row})").font = total_font
        ws.cell(row=total_row, column=2).number_format = "#,##0"

        ws.cell(row=total_row, column=4, value=f"=SUM(D10:D{end_row})").font = total_font
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=4).number_format = "#,##0"

        ws.cell(row=total_row, column=5, value=f"=SUM(E10:E{end_row})").font = total_font
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=5).number_format = "₹#,##0.00"

        for c in range(1, 7):
            cell = ws.cell(row=total_row, column=c)
            cell.border = border_total
            cell.fill = total_fill

        # Auto filter
        ws.auto_filter.ref = f"A9:F{end_row}"

        cls._auto_adjust_columns(ws)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @classmethod
    def generate_orders_report(cls, start_date: str, end_date: str, orders: list) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"

        (
            title_font, meta_font, section_font, header_font, data_font, total_font,
            header_fill, total_fill, kpi_fill, border_data, border_total, border_kpi
        ) = cls._apply_styles(ws, 12)

        # Title
        ws["A1"] = "CHOVIQUE LUXURY CHOCOLATES — ORDERS REPORT"
        ws["A1"].font = title_font
        ws.merge_cells("A1:L1")

        # Date Range
        ws["A2"] = f"Date Range: {start_date} to {end_date}"
        ws["A2"].font = meta_font
        ws.merge_cells("A2:L2")

        # Table Header
        ws["A4"] = "DETAILED ORDERS DATA"
        ws["A4"].font = section_font

        headers = [
            "Order ID", "Customer Name", "Customer Email", "Order Date", "Number of Items",
            "Subtotal", "Discount", "Shipping", "Tax", "Total Amount", "Payment Status", "Order Status"
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left" if col_idx in [1, 2, 3, 11, 12] else "right")

        # Freeze panes
        ws.freeze_panes = "A6"

        # Data rows
        start_row = 6
        for r_idx, ord_data in enumerate(orders, start_row):
            # Order ID, Customer Name, Email
            ws.cell(row=r_idx, column=1, value=ord_data[0]).font = data_font
            ws.cell(row=r_idx, column=2, value=ord_data[1]).font = data_font
            ws.cell(row=r_idx, column=3, value=ord_data[2]).font = data_font

            # Order Date
            date_cell = ws.cell(row=r_idx, column=4)
            try:
                date_cell.value = datetime.strptime(ord_data[3], "%Y-%m-%d %H:%M")
                date_cell.number_format = "YYYY-MM-DD HH:MM"
            except:
                date_cell.value = ord_data[3]
            date_cell.font = data_font
            date_cell.alignment = Alignment(horizontal="right")

            # Number of Items
            items_cell = ws.cell(row=r_idx, column=5, value=int(ord_data[4] or 0))
            items_cell.font = data_font
            items_cell.alignment = Alignment(horizontal="right")
            items_cell.number_format = "#,##0"

            # Subtotal, Discount, Shipping, Tax, Total Amount
            for col_pos, val in enumerate(ord_data[5:10], 6):
                # Clean currency string to float
                if isinstance(val, str):
                    val = float(val.replace("₹", "").replace(",", ""))
                val_cell = ws.cell(row=r_idx, column=col_pos, value=float(val or 0.0))
                val_cell.font = data_font
                val_cell.alignment = Alignment(horizontal="right")
                val_cell.number_format = "₹#,##0.00"

            # Payment Status, Order Status
            ws.cell(row=r_idx, column=11, value=ord_data[10]).font = data_font
            ws.cell(row=r_idx, column=12, value=ord_data[11]).font = data_font

            for c in range(1, 13):
                ws.cell(row=r_idx, column=c).border = border_data

        end_row = start_row + len(orders) - 1
        total_row = end_row + 1

        # Totals Row
        ws.cell(row=total_row, column=1, value="Total").font = total_font
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")

        # Total Orders Count
        ws.cell(row=total_row, column=2, value=f"=COUNTA(A6:A{end_row})").font = total_font
        ws.cell(row=total_row, column=2).number_format = "#,##0"

        # Total Items
        ws.cell(row=total_row, column=5, value=f"=SUM(E6:E{end_row})").font = total_font
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=5).number_format = "#,##0"

        # Total Amount Totals (Subtotal to Total Amount)
        for col_pos in range(6, 11):
            col_letter = get_column_letter(col_pos)
            tot_cell = ws.cell(row=total_row, column=col_pos, value=f"=SUM({col_letter}6:{col_letter}{end_row})")
            tot_cell.font = total_font
            tot_cell.alignment = Alignment(horizontal="right")
            tot_cell.number_format = "₹#,##0.00"

        for c in range(1, 13):
            cell = ws.cell(row=total_row, column=c)
            cell.border = border_total
            cell.fill = total_fill

        ws.auto_filter.ref = f"A5:L{end_row}"
        cls._auto_adjust_columns(ws)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @classmethod
    def generate_products_report(cls, start_date: str, end_date: str, products: list) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Report"

        (
            title_font, meta_font, section_font, header_font, data_font, total_font,
            header_fill, total_fill, kpi_fill, border_data, border_total, border_kpi
        ) = cls._apply_styles(ws, 7)

        # Title
        ws["A1"] = "CHOVIQUE LUXURY CHOCOLATES — PRODUCTS REPORT"
        ws["A1"].font = title_font
        ws.merge_cells("A1:G1")

        # Date Range
        ws["A2"] = f"Date Range: {start_date} to {end_date}"
        ws["A2"].font = meta_font
        ws.merge_cells("A2:G2")

        # Table Header
        ws["A4"] = "PRODUCT PERFORMANCE SUMMARY"
        ws["A4"].font = section_font

        headers = ["Product Name", "Category", "Units Sold", "Total Orders", "Revenue", "Average Selling Price", "Stock Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left" if col_idx in [1, 2, 7] else "right")

        ws.freeze_panes = "A6"

        start_row = 6
        for r_idx, prod in enumerate(products, start_row):
            # Name, Category
            ws.cell(row=r_idx, column=1, value=prod[0]).font = data_font
            ws.cell(row=r_idx, column=2, value=prod[1]).font = data_font

            # Units Sold
            units = int(str(prod[2]).replace(",", "")) if prod[2] else 0
            units_cell = ws.cell(row=r_idx, column=3, value=units)
            units_cell.font = data_font
            units_cell.alignment = Alignment(horizontal="right")
            units_cell.number_format = "#,##0"

            # Total Orders
            total_ords = int(prod[3]) if len(prod) > 6 and prod[3] is not None else int(units)  # Fallback
            orders_cell = ws.cell(row=r_idx, column=4, value=total_ords)
            orders_cell.font = data_font
            orders_cell.alignment = Alignment(horizontal="right")
            orders_cell.number_format = "#,##0"

            # Revenue
            rev_str = str(prod[4]) if len(prod) > 4 else str(prod[2])
            rev_val = float(rev_str.replace("₹", "").replace(",", "")) if rev_str else 0.0
            rev_cell = ws.cell(row=r_idx, column=5, value=rev_val)
            rev_cell.font = data_font
            rev_cell.alignment = Alignment(horizontal="right")
            rev_cell.number_format = "₹#,##0.00"

            # Average Selling Price
            avg_price = rev_val / units if units > 0 else 0.0
            asp_cell = ws.cell(row=r_idx, column=6, value=avg_price)
            asp_cell.font = data_font
            asp_cell.alignment = Alignment(horizontal="right")
            asp_cell.number_format = "₹#,##0.00"

            # Stock Status / Stock level
            stock_lvl = prod[3] if len(prod) <= 6 else prod[6] # From list representation
            ws.cell(row=r_idx, column=7, value=f"{stock_lvl} in stock" if isinstance(stock_lvl, int) else str(stock_lvl)).font = data_font

            for c in range(1, 8):
                ws.cell(row=r_idx, column=c).border = border_data

        end_row = start_row + len(products) - 1
        total_row = end_row + 1

        # Totals Row
        ws.cell(row=total_row, column=1, value="Total").font = total_font
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")

        # Total Units Sold
        ws.cell(row=total_row, column=3, value=f"=SUM(C6:C{end_row})").font = total_font
        ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=3).number_format = "#,##0"

        # Total Orders
        ws.cell(row=total_row, column=4, value=f"=SUM(D6:D{end_row})").font = total_font
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=4).number_format = "#,##0"

        # Total Revenue
        ws.cell(row=total_row, column=5, value=f"=SUM(E6:E{end_row})").font = total_font
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=5).number_format = "₹#,##0.00"

        for c in range(1, 8):
            cell = ws.cell(row=total_row, column=c)
            cell.border = border_total
            cell.fill = total_fill

        ws.auto_filter.ref = f"A5:G{end_row}"
        cls._auto_adjust_columns(ws)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @classmethod
    def generate_analytics_report(cls, start_date: str, end_date: str, summary_data: dict) -> io.BytesIO:
        wb = Workbook()
        
        # Sheet 1: Summary Sheet
        ws = wb.active
        ws.title = "Analytics Summary"

        (
            title_font, meta_font, section_font, header_font, data_font, total_font,
            header_fill, total_fill, kpi_fill, border_data, border_total, border_kpi
        ) = cls._apply_styles(ws, 3)

        # Title
        ws["A1"] = "CHOVIQUE LUXURY CHOCOLATES — PERFORMANCE SUMMARY"
        ws["A1"].font = title_font
        ws.merge_cells("A1:C1")

        # Date Range
        ws["A2"] = f"Date Range: {start_date} to {end_date}"
        ws["A2"].font = meta_font
        ws.merge_cells("A2:C2")

        # Summary KPIs Header
        ws["A4"] = "METRICS SUMMARY"
        ws["A4"].font = section_font

        headers = ["Metric Key", "Metric Value", "Descriptor"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left" if col_idx != 2 else "right")

        metrics = [
            ("Total Revenue", summary_data.get("total_revenue", 0.0), "₹#,##0.00"),
            ("Total Orders", summary_data.get("total_orders", 0), "#,##0"),
            ("Total Customers", summary_data.get("total_customers", 0), "#,##0"),
            ("New Customers", summary_data.get("new_customers", 0), "#,##0"),
            ("Repeat Customers", summary_data.get("repeat_customers", 0), "#,##0"),
            ("Average Order Value", summary_data.get("avg_order_value", 0.0), "₹#,##0.00"),
            ("Total Products Sold", summary_data.get("total_products_sold", 0), "#,##0"),
            ("Total Discounts", summary_data.get("total_discounts", 0.0), "₹#,##0.00"),
            ("Total Tax", summary_data.get("total_tax", 0.0), "₹#,##0.00"),
            ("Total Shipping Revenue", summary_data.get("total_shipping_revenue", 0.0), "₹#,##0.00"),
        ]

        for r_idx, (m_name, m_val, m_fmt) in enumerate(metrics, 6):
            ws.cell(row=r_idx, column=1, value=m_name).font = data_font
            
            val_cell = ws.cell(row=r_idx, column=2, value=float(m_val) if isinstance(m_val, (int, float)) else m_val)
            val_cell.font = total_font
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.number_format = m_fmt

            ws.cell(row=r_idx, column=3, value="Cumulative value").font = meta_font

            for c in range(1, 4):
                ws.cell(row=r_idx, column=c).border = border_data

        cls._auto_adjust_columns(ws)

        # Add Daily Trend sheet if data is present
        if "daily_trend" in summary_data and summary_data["daily_trend"]:
            ws2 = wb.create_sheet(title="Daily Sales Trend")
            ws2.views.sheetView[0].showGridLines = True
            
            ws2["A1"] = "DAILY SALES REVENUE & VOLUME TREND"
            ws2["A1"].font = title_font
            ws2.merge_cells("A1:D1")

            ws2.cell(row=3, column=1, value="Date").font = header_font
            ws2.cell(row=3, column=1).fill = header_fill
            
            ws2.cell(row=3, column=2, value="Orders Count").font = header_font
            ws2.cell(row=3, column=2).fill = header_fill
            ws2.cell(row=3, column=2).alignment = Alignment(horizontal="right")

            ws2.cell(row=3, column=3, value="Total Revenue").font = header_font
            ws2.cell(row=3, column=3).fill = header_fill
            ws2.cell(row=3, column=3).alignment = Alignment(horizontal="right")

            ws2.freeze_panes = "A4"

            for tr_idx, day_val in enumerate(summary_data["daily_trend"], 4):
                ws2.cell(row=tr_idx, column=1, value=day_val[0]).font = data_font
                
                cnt_cell = ws2.cell(row=tr_idx, column=2, value=int(day_val[1]))
                cnt_cell.font = data_font
                cnt_cell.alignment = Alignment(horizontal="right")
                cnt_cell.number_format = "#,##0"

                rev_cell = ws2.cell(row=tr_idx, column=3, value=float(day_val[2]))
                rev_cell.font = data_font
                rev_cell.alignment = Alignment(horizontal="right")
                rev_cell.number_format = "₹#,##0.00"

                for col_idx in range(1, 4):
                    ws2.cell(row=tr_idx, column=col_idx).border = border_data

            cls._auto_adjust_columns(ws2)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out
